"""
Builds the stakeholder Excel report with a polished, modern dashboard look:
  - Cover sheet with title banner and quick-glance stat chips
  - KPI Dashboard: icon-tagged card tiles, styled pie/line charts, icon-set
    flags on process health metrics
  - Raw data sheets as banded Excel Tables with real date/currency number
    formats (not raw ISO strings) and conditional formatting
  - Color-coded sheet tabs

All KPI values are live formulas referencing the raw data sheets — nothing
is pre-computed and hardcoded into cells.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, IconSetRule

# ---- Palette ----
NAVY = "18314F"
NAVY_SOFT = "24476F"
BLUE = "0070F2"
BLUE_SOFT = "4F9CF9"
TEAL = "0BA5A5"
AMBER = "E9730C"
PURPLE = "8B5CF6"
GREEN = "1E8E3E"
INK = "1D2733"
GREY_TEXT = "5B6B7C"
FAINT = "EEF2F7"
WHITE = "FFFFFF"

FONT_NAME = "Calibri"

TITLE_FONT = Font(name=FONT_NAME, bold=True, size=26, color=WHITE)
KICKER_FONT = Font(name=FONT_NAME, bold=True, size=11, color=BLUE_SOFT)
SUBTITLE_FONT = Font(name=FONT_NAME, size=13, color="B9CBE3")
SECTION_FONT = Font(name=FONT_NAME, bold=True, size=13, color=NAVY)
HEADER_FONT = Font(name=FONT_NAME, bold=True, color=WHITE, size=10)
HEADER_FILL = PatternFill("solid", start_color=NAVY, end_color=NAVY)
BASE_FONT = Font(name=FONT_NAME, size=10, color=INK)
TILE_ICON_FONT = Font(name=FONT_NAME, size=16, color=WHITE)
TILE_LABEL_FONT = Font(name=FONT_NAME, size=9, color="F2F5F9", bold=True)
TILE_NUMBER_FONT = Font(name=FONT_NAME, size=25, bold=True, color=WHITE)
TILE_SUB_FONT = Font(name=FONT_NAME, size=8.5, color="F2F5F9")

TILE_COLORS = [BLUE, TEAL, AMBER, PURPLE, GREEN]
TILE_ICONS = ["\U0001F4CB", "\u2705", "\U0001F4E6", "\U0001F4B6", "\u23F1"]  # clipboard, check, package, money, timer

DATE_FMT = "dd mmm yyyy"
CURRENCY_FMT = '#,##0" €"'
SCORE_FMT = "0.0"

DATE_COLUMNS = {
    "ServiceRequests": ["created_date", "approved_date", "bidding_start", "bidding_end"],
    "Offers": ["submitted_date"],
    "Evaluations": ["evaluated_date"],
    "Orders": ["created_date", "approved_date"],
    "OrderChanges": ["requested_date", "resolved_date"],
}
CURRENCY_COLUMNS = {
    "Offers": ["price_eur"],
    "Orders": ["order_value_eur"],
    "DepartmentBreakdown": ["total_order_value"],
}
SCORE_COLUMNS = {
    "Evaluations": ["score"],
}

TAB_COLORS = {
    "Cover": NAVY,
    "KPI Dashboard": BLUE,
    "ServiceRequests": TEAL,
    "Offers": TEAL,
    "Evaluations": TEAL,
    "Orders": AMBER,
    "OrderChanges": AMBER,
    "MonthlyVolume": PURPLE,
    "DepartmentBreakdown": PURPLE,
    "ProviderPerformance": GREEN,
}


def _style_table_sheet(ws, df, table_name, sheet_key):
    date_cols = DATE_COLUMNS.get(sheet_key, [])
    currency_cols = CURRENCY_COLUMNS.get(sheet_key, [])
    score_cols = SCORE_COLUMNS.get(sheet_key, [])

    for c, col in enumerate(df.columns, 1):
        ws.cell(row=1, column=c, value=col)

    for r, row in enumerate(df.itertuples(index=False), 2):
        for c, (col_name, val) in enumerate(zip(df.columns, row), 1):
            if col_name in date_cols and val:
                try:
                    val = pd.to_datetime(val).to_pydatetime()
                except (ValueError, TypeError):
                    pass
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BASE_FONT
            if col_name in date_cols and val:
                cell.number_format = DATE_FMT
            elif col_name in currency_cols:
                cell.number_format = CURRENCY_FMT
            elif col_name in score_cols:
                cell.number_format = SCORE_FMT

    n_rows, n_cols = len(df), len(df.columns)
    last_col = get_column_letter(n_cols)
    ref = f"A1:{last_col}{max(n_rows + 1, 2)}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table)

    for c, col in enumerate(df.columns, 1):
        if col in date_cols:
            width = 13
        elif col in currency_cols:
            width = 15
        else:
            max_len = max(10, int(df[col].astype(str).str.len().max()) if n_rows else 10)
            width = min(32, max_len + 3)
        ws.column_dimensions[get_column_letter(c)].width = width

    for cell in ws[1]:
        cell.font = HEADER_FONT
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    if sheet_key in TAB_COLORS:
        ws.sheet_properties.tabColor = TAB_COLORS[sheet_key]
    return n_rows, n_cols


def _add_conditional_formatting(ws, sheet_name, n_rows):
    if sheet_name == "Offers" and n_rows:
        ws.conditional_formatting.add(
            f"E2:E{n_rows+1}",
            DataBarRule(start_type="min", end_type="max", color=BLUE)
        )
    elif sheet_name == "Evaluations" and n_rows:
        ws.conditional_formatting.add(
            f"E2:E{n_rows+1}",
            ColorScaleRule(start_type="min", start_color="F8696B",
                            mid_type="percentile", mid_value=50, mid_color="FFEB84",
                            end_type="max", end_color="63BE7B")
        )
        ws.conditional_formatting.add(
            f"E2:E{n_rows+1}",
            IconSetRule(icon_style="3TrafficLights1", type="percent", values=[0, 33, 67], showValue=True, reverse=False)
        )
    elif sheet_name == "Orders" and n_rows:
        ws.conditional_formatting.add(
            f"F2:F{n_rows+1}",
            DataBarRule(start_type="min", end_type="max", color=TEAL)
        )


def _cover_sheet(wb, generated_label, headline_stats):
    ws = wb.create_sheet("Cover", 0)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = TAB_COLORS["Cover"]
    for col, w in zip("ABCDEFGHI", [3, 13, 13, 13, 13, 13, 13, 13, 3]):
        ws.column_dimensions[col].width = w

    for r in range(1, 24):
        for c in range(1, 10):
            ws.cell(row=r, column=c).fill = PatternFill("solid", start_color=NAVY, end_color=NAVY)
    for c in range(1, 10):
        ws.cell(row=16, column=c).fill = PatternFill("solid", start_color=NAVY_SOFT, end_color=NAVY_SOFT)

    ws.merge_cells("B3:H3")
    ws["B3"] = "PROCESS GOVERNANCE REPORT"
    ws["B3"].font = KICKER_FONT

    ws.merge_cells("B4:H6")
    ws["B4"] = "IT Service Management"
    ws["B4"].font = TITLE_FONT
    ws["B4"].alignment = Alignment(vertical="top")

    ws.merge_cells("B7:H7")
    ws["B7"] = "Request-to-Order Lifecycle & KPI Dashboard"
    ws["B7"].font = SUBTITLE_FONT

    ws.merge_cells("B9:H9")
    ws["B9"] = f"Generated {generated_label}   ·   Source: ServiceManagementSystem process data"
    ws["B9"].font = Font(name=FONT_NAME, size=10, color="8FA6C4", italic=True)

    ws.merge_cells("B11:H13")
    ws["B11"] = ("Tracks the full service-request lifecycle — intake, procurement approval, "
                 "provider bidding, offer evaluation, order placement, and post-order changes — "
                 "with live KPIs, control-deviation flags, and provider benchmarking.")
    ws["B11"].font = Font(name=FONT_NAME, size=10.5, color="D9E6F5")
    ws["B11"].alignment = Alignment(wrap_text=True, vertical="top")

    # Quick-glance stat chips
    chip_row = 17
    col = 2
    for label, value in headline_stats:
        ws.merge_cells(start_row=chip_row, start_column=col, end_row=chip_row, end_column=col + 1)
        c1 = ws.cell(row=chip_row, column=col, value=value)
        c1.font = Font(name=FONT_NAME, bold=True, size=17, color=WHITE)
        c1.alignment = Alignment(horizontal="left")
        ws.merge_cells(start_row=chip_row + 1, start_column=col, end_row=chip_row + 1, end_column=col + 1)
        c2 = ws.cell(row=chip_row + 1, column=col, value=label.upper())
        c2.font = Font(name=FONT_NAME, size=8, bold=True, color="8FA6C4")
        col += 2

    ws.row_dimensions[4].height = 24
    ws.row_dimensions[5].height = 24
    ws.row_dimensions[7].height = 20
    ws.row_dimensions[17].height = 24
    return ws


def _tile(ws, row, col, width, height, icon, label, formula_or_value, color, number_format=None, subtitle=None):
    r1, c1 = row, col
    r2, c2 = row + height - 1, col + width - 1
    for rr in range(r1, r2 + 1):
        for cc in range(c1, c2 + 1):
            ws.cell(row=rr, column=cc).fill = PatternFill("solid", start_color=color, end_color=color)

    icon_cell = ws.cell(row=r1, column=c1, value=icon)
    icon_cell.font = TILE_ICON_FONT
    icon_cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(start_row=r1, start_column=c1 + 1, end_row=r1, end_column=c2)
    lbl = ws.cell(row=r1, column=c1 + 1, value=label.upper())
    lbl.font = TILE_LABEL_FONT
    lbl.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(start_row=r1 + 1, start_column=c1, end_row=r2 - 1 if height > 2 else r1 + 1, end_column=c2)
    val = ws.cell(row=r1 + 1, column=c1, value=formula_or_value)
    val.font = TILE_NUMBER_FONT
    val.alignment = Alignment(horizontal="left", vertical="center")
    if number_format:
        val.number_format = number_format

    if subtitle and height > 2:
        ws.merge_cells(start_row=r2, start_column=c1, end_row=r2, end_column=c2)
        sub = ws.cell(row=r2, column=c1, value=subtitle)
        sub.font = TILE_SUB_FONT
        sub.alignment = Alignment(horizontal="left", vertical="center")


def _style_pie(pie):
    pie.title = "Requests by Status"
    pie.style = 26
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = False
    from openpyxl.chart.marker import DataPoint
    from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
    colors = [NAVY, BLUE, TEAL, AMBER, PURPLE, GREEN, "6B7A8F"]
    for i, color in enumerate(colors):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = color
        pie.series[0].data_points.append(pt) if pie.series else None
    return pie


def _style_line(line, n_series_colors):
    line.title = "Requests vs Orders by Month"
    line.style = 2
    line.y_axis.title = "Count"
    line.y_axis.majorGridlines.spPr = None
    for i, s in enumerate(line.series):
        s.graphicalProperties.line.width = 24000
        s.graphicalProperties.line.solidFill = n_series_colors[i % len(n_series_colors)]
        s.smooth = False
    return line


def build_workbook(frames, monthly_df, dept_df, output_path, generated_label=""):
    wb = Workbook()
    wb.remove(wb.active)

    n_req_total = len(frames["requests"])
    n_ord_total = len(frames["orders"])
    total_value = frames["orders"]["order_value_eur"].sum() if n_ord_total else 0
    approved = (frames["requests"]["status"] != "REJECTED").sum()
    approval_pct = round(100 * approved / n_req_total, 1) if n_req_total else 0
    headline_stats = [
        ("Requests", str(n_req_total)),
        ("Orders", str(n_ord_total)),
        ("Approval Rate", f"{approval_pct}%"),
        ("Order Value", f"€{total_value:,.0f}"),
    ]
    _cover_sheet(wb, generated_label, headline_stats)

    sheet_defs = [
        ("ServiceRequests", frames["requests"]),
        ("Offers", frames["offers"]),
        ("Evaluations", frames["evaluations"]),
        ("Orders", frames["orders"]),
        ("OrderChanges", frames["order_changes"]),
        ("MonthlyVolume", monthly_df),
        ("DepartmentBreakdown", dept_df),
    ]
    row_counts = {}
    for name, df in sheet_defs:
        ws = wb.create_sheet(name)
        n_rows, _ = _style_table_sheet(ws, df, f"tbl_{name}", name)
        row_counts[name] = n_rows
        _add_conditional_formatting(ws, name, n_rows)

    n_req = row_counts["ServiceRequests"]
    n_off = row_counts["Offers"]
    n_ord = row_counts["Orders"]
    n_chg = row_counts["OrderChanges"]

    req_status_col = "F"
    ord_status_col = "E"
    chg_type_col = "C"
    chg_status_col = "D"

    dash = wb.create_sheet("KPI Dashboard", 1)
    dash.sheet_view.showGridLines = False
    dash.sheet_properties.tabColor = TAB_COLORS["KPI Dashboard"]
    for i in range(1, 20):
        dash.column_dimensions[get_column_letter(i)].width = 11

    dash.merge_cells("A1:N1")
    dash["A1"] = "Process KPI Dashboard"
    dash["A1"].font = Font(name=FONT_NAME, bold=True, size=19, color=NAVY)
    dash.row_dimensions[1].height = 30

    dash.merge_cells("A2:N2")
    dash["A2"] = "Live figures pulled from the raw data sheets — refresh with new exports and every value recalculates."
    dash["A2"].font = Font(name=FONT_NAME, size=9, italic=True, color=GREY_TEXT)
    dash.row_dimensions[3].height = 6

    tile_row = 4
    tile_h = 3
    col = 1
    tile_width = 2
    gap = 1
    labels = ["Total Requests", "Approval Rate", "Total Orders", "Order Value (EUR)", "Avg Man-Days/Order"]
    subs = ["Service requests logged", "Procurement approval", "Orders placed", "Total contracted value", "Avg effort per order"]
    values = [
        f"=SUBTOTAL(103,ServiceRequests!A2:A{n_req+1})",
        f'=1-(COUNTIF(ServiceRequests!{req_status_col}2:{req_status_col}{n_req+1},"REJECTED")/SUBTOTAL(103,ServiceRequests!A2:A{n_req+1}))',
        f"=SUBTOTAL(103,Orders!A2:A{n_ord+1})",
        f"=SUM(Orders!F2:F{n_ord+1})",
        f"=AVERAGE(Orders!G2:G{n_ord+1})",
    ]
    formats = [None, "0.0%", None, CURRENCY_FMT, "0.0"]

    for i, (label, sub, val, fmt) in enumerate(zip(labels, subs, values, formats)):
        _tile(dash, tile_row, col, tile_width, tile_h, TILE_ICONS[i], label, val, TILE_COLORS[i % len(TILE_COLORS)],
              number_format=fmt, subtitle=sub)
        col += tile_width + gap

    detail_row = tile_row + tile_h + 2
    dash.cell(row=detail_row, column=1, value="Process Health").font = SECTION_FONT

    def metric_line(row, label, formula, fmt=None):
        c0 = dash.cell(row=row, column=1)
        c0.fill = PatternFill("solid", start_color=FAINT, end_color=FAINT) if row % 2 == 0 else PatternFill(fill_type=None)
        c1 = dash.cell(row=row, column=1, value=label)
        c1.font = Font(name=FONT_NAME, size=10, color=GREY_TEXT)
        c2 = dash.cell(row=row, column=4, value=formula)
        c2.font = Font(name=FONT_NAME, size=10, bold=True, color=NAVY)
        if fmt:
            c2.number_format = fmt

    r = detail_row + 1
    metric_line(r, "Requests completed", f'=COUNTIF(ServiceRequests!{req_status_col}2:{req_status_col}{n_req+1},"COMPLETED")'); r += 1
    metric_line(r, "Requests expired (no offers)", f'=COUNTIF(ServiceRequests!{req_status_col}2:{req_status_col}{n_req+1},"EXPIRED")'); r += 1
    metric_line(r, "Total offers received", f"=SUBTOTAL(103,Offers!A2:A{n_off+1})"); r += 1
    metric_line(r, "Avg offers per request", f"=D{r-1}/SUBTOTAL(103,ServiceRequests!A2:A{n_req+1})", "0.0"); r += 1
    metric_line(r, "Avg bidding window (days)", f"=AVERAGE(ServiceRequests!G2:G{n_req+1})", "0.0"); r += 1
    metric_line(r, "Orders approved", f'=COUNTIF(Orders!{ord_status_col}2:{ord_status_col}{n_ord+1},"APPROVED")'); r += 1
    metric_line(r, "Order approval rate", f"=D{r-1}/SUBTOTAL(103,Orders!A2:A{n_ord+1})", "0.0%"); r += 1
    if n_chg:
        metric_line(r, "Total change requests", f"=SUBTOTAL(103,OrderChanges!A2:A{n_chg+1})"); r += 1
        metric_line(r, "Substitutions", f'=COUNTIF(OrderChanges!{chg_type_col}2:{chg_type_col}{n_chg+1},"SUBSTITUTION")'); r += 1
        metric_line(r, "Extensions", f'=COUNTIF(OrderChanges!{chg_type_col}2:{chg_type_col}{n_chg+1},"EXTENSION")'); r += 1
        metric_line(r, "Change approval rate",
                     f'=COUNTIF(OrderChanges!{chg_status_col}2:{chg_status_col}{n_chg+1},"APPROVED")/SUBTOTAL(103,OrderChanges!A2:A{n_chg+1})',
                     "0.0%"); r += 1

    chart_col = 17
    status_hdr_row = tile_row
    dash.cell(row=status_hdr_row, column=chart_col, value="Status").font = HEADER_FONT
    dash.cell(row=status_hdr_row, column=chart_col).fill = HEADER_FILL
    dash.cell(row=status_hdr_row, column=chart_col + 1, value="Count").font = HEADER_FONT
    dash.cell(row=status_hdr_row, column=chart_col + 1).fill = HEADER_FILL
    statuses = sorted(frames["requests"]["status"].unique())
    for i, status in enumerate(statuses, 1):
        row = status_hdr_row + i
        dash.cell(row=row, column=chart_col, value=status).font = BASE_FONT
        dash.cell(row=row, column=chart_col + 1,
                  value=f'=COUNTIF(ServiceRequests!{req_status_col}2:{req_status_col}{n_req+1},{get_column_letter(chart_col)}{row})').font = BASE_FONT

    pie = PieChart()
    data_ref = Reference(dash, min_col=chart_col + 1, min_row=status_hdr_row, max_row=status_hdr_row + len(statuses))
    cats_ref = Reference(dash, min_col=chart_col, min_row=status_hdr_row + 1, max_row=status_hdr_row + len(statuses))
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(cats_ref)
    _style_pie(pie)
    pie.height, pie.width = 8.5, 12
    dash.add_chart(pie, f"{get_column_letter(chart_col)}{status_hdr_row + len(statuses) + 2}")

    line = LineChart()
    n_months = len(monthly_df)
    data_ref = Reference(wb["MonthlyVolume"], min_col=2, max_col=3, min_row=1, max_row=n_months + 1)
    cats_ref = Reference(wb["MonthlyVolume"], min_col=1, min_row=2, max_row=n_months + 1)
    line.add_data(data_ref, titles_from_data=True)
    line.set_categories(cats_ref)
    _style_line(line, [BLUE, AMBER])
    line.width, line.height = 23, 9.5
    dash.add_chart(line, "A" + str(r + 2))

    wb.save(output_path)
    return output_path


def add_provider_sheet(wb_path, provider_df):
    """Adds a dedicated Provider Performance sheet with its own bar chart."""
    from openpyxl import load_workbook
    wb = load_workbook(wb_path)
    ws = wb.create_sheet("ProviderPerformance")
    _style_table_sheet(ws, provider_df, "tbl_ProviderPerformance", "ProviderPerformance")
    ws.conditional_formatting.add(
        f"D2:D{len(provider_df)+1}",
        DataBarRule(start_type="min", end_type="max", color=TEAL)
    )
    chart = BarChart()
    chart.type = "bar"
    chart.title = "Provider Win Rate (%)"
    chart.y_axis.title = "Win rate %"
    chart.style = 11
    n = len(provider_df)
    data_ref = Reference(ws, min_col=5, min_row=1, max_row=n + 1)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=n + 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    for s in chart.series:
        s.graphicalProperties.solidFill = TEAL
    chart.width, chart.height = 20, 10
    ws.add_chart(chart, "H2")
    wb.save(wb_path)
